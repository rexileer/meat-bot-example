from decimal import Decimal
from io import BytesIO

from django.db.models import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from Web.CRM.models import RawMeatBatch
from bot.utils.excel import beautify_columns


async def generate_storage_datatable(meat_batches: QuerySet[RawMeatBatch]) -> BytesIO:

    meat_batches = meat_batches.order_by("created_at").order_by("raw_material__name")

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ID партии",
            "Вид сырья",
            "Общая масса (кг)",
            "Массовая доля жира",
            "Массовая доля белка",
            "Массовая доля влаги",
            "Есть на складе",
        ]
    )
    for meat_batch in meat_batches:
        fat_proportion = ""
        protein_proportion = ""
        moisture_proportion = ""
        meat_batch: RawMeatBatch = meat_batch
        analyze_status = await meat_batch.statuses.filter(status__codename="laboratory_analyze").afirst()
        if analyze_status:
            if analyze_status.additional_data.get("fat_proportion"):
                fat_proportion = Decimal(analyze_status.additional_data["fat_proportion"])
            if analyze_status.additional_data.get("protein_proportion"):
                protein_proportion = Decimal(analyze_status.additional_data["protein_proportion"])
            if analyze_status.additional_data.get("moisture_proportion"):
                moisture_proportion = Decimal(analyze_status.additional_data["moisture_proportion"])
        ws.append(
            (
                f"{meat_batch.production_id} - {meat_batch.organization}",
                meat_batch.raw_material.name,
                meat_batch.weight,
                fat_proportion,
                protein_proportion,
                moisture_proportion,
                "В поставке" if meat_batch.is_future_batch else "Да",
            )
        )

    beautify_columns(ws)

    color_font = Font(size="12", color="FF0000")

    last_name = ""
    for cell in ws["B"][1:]:
        if cell.value != last_name:
            cell.font = color_font
            ws[f"A{cell.row}"].font = color_font
        last_name = cell.value

    return BytesIO(save_virtual_workbook(wb))
