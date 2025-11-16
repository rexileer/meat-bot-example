import logging
from decimal import Decimal
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from Web.CRM.models import (
    RawMeatBatch,
    MeatBlank,
    SecondMincedMeat,
    MeatBlankRawMeatBatch,
    RawMeatBatchStatus,
    MincedMeatBatchSecondMeatBlank,
)
from bot.utils.excel import beautify_columns


async def generate_storage_datatable(meat_batches: List[RawMeatBatch]) -> BytesIO:
    meat_batches.sort(key=lambda item: (item.raw_material.name, item.created_at))

    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ID партии",
            "Вид сырья",
            "Общая масса (кг)",
            "Массовая доля жира",
            "Массовая доля белка",
            "Массовая доля влаги",
            "Есть на складе",
        ]
    )
    for meat_batch in meat_batches:
        fat_proportion = ""
        protein_proportion = ""
        moisture_proportion = ""

        analyze_status = await meat_batch.statuses.filter(status__codename="laboratory_analyze").afirst()
        if analyze_status:
            if analyze_status.additional_data.get("fat_proportion"):
                fat_proportion = Decimal(analyze_status.additional_data["fat_proportion"])
            if analyze_status.additional_data.get("protein_proportion"):
                protein_proportion = Decimal(analyze_status.additional_data["protein_proportion"])
            if analyze_status.additional_data.get("moisture_proportion"):
                moisture_proportion = Decimal(analyze_status.additional_data["moisture_proportion"])
        ws.append(
            (
                f"{meat_batch.production_id} - {meat_batch.organization}",
                meat_batch.raw_material.name,
                meat_batch.weight,
                fat_proportion,
                protein_proportion,
                moisture_proportion,
                "В поставке" if meat_batch.is_future_batch else "Да",
            )
        )

    beautify_columns(ws)

    color_font = Font(size="12", color="FF0000")

    last_name = ""
    for cell in ws["B"][1:]:
        if cell.value != last_name:
            cell.font = color_font
            ws[f"A{cell.row}"].font = color_font
        last_name = cell.value

    return BytesIO(save_virtual_workbook(wb))


async def generate_blanks_datatable(meat_blanks: List[MeatBlank], chilled_meat_batches: List[RawMeatBatch]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ID заготовки",
            "ID партии",
            "Вид сырья",
            "Общая масса (кг)",
            "Массовая доля жира",
            "Массовая доля белка",
            "Массовая доля влаги",
            "Есть на складе",
        ]
    )
    logging.info(meat_blanks)
    for meat_blank in meat_blanks:
        ws.append((meat_blank.production_id,))
        meat_batches = [
            i.raw_meat_batch
            for i in MeatBlankRawMeatBatch.objects.filter(meat_blank__production_id=meat_blank.production_id).all()
        ]
        meat_batches.sort(key=lambda item: (item.raw_material.name, item.created_at))

        for meat_batch in meat_batches:
            fat_proportion = ""
            protein_proportion = ""
            moisture_proportion = ""

            analyze_status = await meat_batch.statuses.filter(status__codename="laboratory_analyze").afirst()

            if analyze_status:
                if analyze_status.additional_data.get("fat_proportion"):
                    fat_proportion = Decimal(analyze_status.additional_data["fat_proportion"])
                if analyze_status.additional_data.get("protein_proportion"):
                    protein_proportion = Decimal(analyze_status.additional_data["protein_proportion"])
                if analyze_status.additional_data.get("moisture_proportion"):
                    moisture_proportion = Decimal(analyze_status.additional_data["moisture_proportion"])
            ws.append(
                (
                    "",
                    f"{meat_batch.production_id} - {meat_batch.organization}",
                    meat_batch.raw_material.name,
                    meat_blank.weight,
                    fat_proportion,
                    protein_proportion,
                    moisture_proportion,
                    "В поставке" if meat_batch.is_future_batch else "Да",
                )
            )
    ws.append(("",))
    ws.append(("Охлажденное сырье",))
    for meat_batch in chilled_meat_batches:
        fat_proportion = ""
        protein_proportion = ""
        moisture_proportion = ""

        analyze_status = await RawMeatBatchStatus.objects.filter(
            status__codename="laboratory_analyze", raw_meat_batch_id=meat_batch.pk
        ).afirst()

        if analyze_status:
            if analyze_status.additional_data.get("fat_proportion"):
                fat_proportion = Decimal(analyze_status.additional_data["fat_proportion"])
            if analyze_status.additional_data.get("protein_proportion"):
                protein_proportion = Decimal(analyze_status.additional_data["protein_proportion"])
            if analyze_status.additional_data.get("moisture_proportion"):
                moisture_proportion = Decimal(analyze_status.additional_data["moisture_proportion"])
        ws.append(
            (
                "",
                f"{meat_batch.production_id} - {meat_batch.organization}",
                meat_batch.raw_material.name,
                meat_batch.weight,
                fat_proportion,
                protein_proportion,
                moisture_proportion,
                "В поставке" if meat_batch.is_future_batch else "Да",
            )
        )

    beautify_columns(ws)

    color_font = Font(size="12", color="FF0000")

    last_name = ""
    for cell in ws["C"][1:]:
        if cell.value != last_name:
            cell.font = color_font
            ws[f"A{cell.row}"].font = color_font
        last_name = cell.value

    return BytesIO(save_virtual_workbook(wb))


async def generate_blanks_datatable_mko(
    meat_blanks: List[SecondMincedMeat], chilled_meat_batches: List[RawMeatBatch]
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "ID заготовки",
            "ID партии",
            "Вид сырья",
            "Общая масса (кг)",
            "Массовая доля жира",
            "Массовая доля белка",
            "Массовая доля влаги",
            "Есть на складе",
        ]
    )
    logging.info(meat_blanks)
    for meat_blank in chilled_meat_batches:
        ws.append((meat_blank.production_id,))
        meat_batches = [
            i.raw_meat_batch
            for i in MeatBlankRawMeatBatch.objects.filter(
                meat_blank__production_id=meat_blank.production_id, meat_blank__type_meat_blank=1
            ).all()
        ]
        meat_batches.sort(key=lambda item: (item.raw_material.name, item.created_at))

        for meat_batch in meat_batches:
            fat_proportion = ""
            protein_proportion = ""
            moisture_proportion = ""
            actual_weight = MeatBlankRawMeatBatch.objects.get(
                meat_blank_id=meat_blank.pk, raw_meat_batch_id=meat_batch.pk
            )

            analyze_status = await meat_batch.statuses.filter(status__codename="laboratory_analyze").afirst()

            if analyze_status:
                if analyze_status.additional_data.get("fat_proportion"):
                    fat_proportion = Decimal(analyze_status.additional_data["fat_proportion"])
                if analyze_status.additional_data.get("protein_proportion"):
                    protein_proportion = Decimal(analyze_status.additional_data["protein_proportion"])
                if analyze_status.additional_data.get("moisture_proportion"):
                    moisture_proportion = Decimal(analyze_status.additional_data["moisture_proportion"])
            ws.append(
                (
                    "",
                    f"{meat_batch.production_id} - {meat_batch.organization}",
                    meat_batch.raw_material.name,
                    actual_weight.weight,
                    fat_proportion,
                    protein_proportion,
                    moisture_proportion,
                    "В поставке" if meat_batch.is_future_batch else "Да",
                )
            )

    ws.append(("",))
    ws.append(("Вторфарш",))
    for meat_blank in meat_blanks:
        used_weight = sum(
            [i.weight for i in MincedMeatBatchSecondMeatBlank.objects.filter(second_minced_meat=meat_blank)]
        )
        type_second_minced_meat = "Вторфарш" if meat_blank.type == 0 else f"МКО{meat_blank.type}"
        ws.append(
            (
                "",
                f"{meat_blank.created_at: %d%m}/{meat_blank.id}/{type_second_minced_meat}",
                "",
                f"{meat_blank.weight-used_weight}",
            )
        )

    beautify_columns(ws)

    color_font = Font(size="12", color="FF0000")

    last_name = ""
    for cell in ws["C"][1:]:
        if cell.value != last_name:
            cell.font = color_font
            ws[f"A{cell.row}"].font = color_font
        last_name = cell.value

    return BytesIO(save_virtual_workbook(wb))
