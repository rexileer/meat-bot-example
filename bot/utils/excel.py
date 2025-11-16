from typing import Optional

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def beautify_columns_(ws: Worksheet) -> Worksheet:
    font_style = Font(size="12")

    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            curr_length = len("" if cell.value is None else str(cell.value))
            if curr_length > length:
                length = curr_length

            cell.alignment = Alignment(horizontal="center")
            cell.font = font_style

        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 7

    return ws


THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)

MEDIUM_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium")
)


def beautify_columns(ws: Worksheet, font_size: int = 12, center: bool = True) -> Worksheet:
    font_style = Font(size=font_size)

    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            curr_length = len("" if cell.value is None else str(cell.value))
            if curr_length > length:
                length = curr_length

            if center:
                cell.alignment = Alignment(horizontal="center")
            cell.font = font_style

        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length * font_size ** (
            font_size * 0.013
        )

    return ws


def set_global_border(ws, border):
    for column_cells in ws.columns:
        for cell in column_cells:
            cell.border = border

    return ws


def set_border(ws, cell_range: str, border: Border):
    rows = ws[cell_range]

    rows = list(rows)
    max_y = len(rows) - 1
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1
        for pos_x, cell in enumerate(cells):
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def set_border_by_range(ws, cell_range: str, border: Optional[Border]):
    rows = ws[cell_range]

    rows = list(rows)
    for cells in rows:
        for cell in cells:
            cell.border = border
