from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional

import openpyxl
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers, PatternFill
from openpyxl.writer.excel import save_virtual_workbook

from Web.CRM.models import (
    MincedMeatBatchMix,
    RawMeatBatch,
    MincedMeatBatchRawMeatBatch,
    Shipment,
    RawMeatBatchStatus,
    ShipmentPallet,
    RawMaterialParams,
    MincedMeatBatchStatus,
    SecondMincedMeat,
)
from bot.utils.excel import (
    beautify_columns,
    set_global_border,
    THIN_BORDER,
    set_border_by_range,
)


def generate_raw_meat_used():
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"
    rows_count = 3
    ws.append(["", "", f"{now():%d.%m.%Y}"])
    ws.append(
        [
            "Производитель",
            "остаток на начало дня",
            "приход",
            "расход",
            "остаток на конец дня",
        ]
    )
    raw_material_list, data = RawMeatBatch.used_and_filtred_raw_meats_today()
    for key, val in data.items():
        # Для партий, созданных сегодня, начало дня не может быть отрицательным
        start_of_day = max(
            0, val["cur_weight"] + val["used_weight"] - val["add_weight"]
        )
        ws.append(
            [
                f"{key.split('---')[0]} {key.split('---')[1]}",
                f"{start_of_day}",
                f"{val['add_weight']}",
                val["used_weight"],
                f"{val['cur_weight']}",
            ]
        )
        rows_count += 1
    ws.append(
        [
            "",
        ]
    )
    ws.append(["Итого:", sum([float(ws[f"E{i}"].value) for i in range(3, rows_count)])])
    for material in raw_material_list:
        weight = 0
        for key, val in data.items():
            weight += val["cur_weight"] if material in key else 0
        ws.append([f"{material}", f"{weight}"])
    beautify_columns(ws)
    return BytesIO(save_virtual_workbook(wb))


def generate_minced_meat_batch_mix_groups() -> dict:
    groups = {}
    start = (datetime.utcnow()).replace(hour=7, minute=0, second=0, microsecond=0)
    end = datetime.utcnow().replace(
        hour=7, minute=0, second=0, microsecond=0
    ) + timedelta(days=1)
    
    # Считаем выпущенными только замесы, поступившие на склад ГП (после принятия)
    # Критерий: наличие статуса palletizing_end, созданного в заданном окне
    released_mixes = (
        MincedMeatBatchMix.objects.filter(
            statuses__status__codename="palletizing_end",
            statuses__created_at__gte=start,
            statuses__created_at__lte=end,
        )
        .exclude(statuses__status__codename="mix_is_blocked_analyze")
        .all()
    )

    for mix in released_mixes:
        recipe_name = mix.minced_meat_batch.recipe.name
        
        # Получаем вес замеса
        weight_raw = 0
        palletizing_status = mix.statuses.filter(status__codename="palletizing_end").first()
        if palletizing_status and palletizing_status.additional_data:
            weight_raw = palletizing_status.additional_data.get("weight_raw", 0)
        else:
            # Фоллбек: если нет данных о весе в palletizing_end, используем вес партии
            weight_raw = mix.minced_meat_batch.weight or 0
        
        if groups.get(recipe_name):
            groups[recipe_name]["weight"] += weight_raw
        else:
            groups.update(
                {
                    recipe_name: {
                        "weight": weight_raw
                    }
                }
            )
    return groups


def generate_meat_batch_datatable() -> BytesIO:
    date_format = "%d.%m.%Y"
    start = (datetime.utcnow()).replace(hour=7, minute=0, second=0, microsecond=0)
    end = datetime.utcnow() + timedelta(days=1)

    raw_material_list, data = RawMeatBatch.used_and_filtred_raw_meats_today()
    minced_meat_batch_groups = generate_minced_meat_batch_mix_groups()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты"

    weight_sum = 0
    minced_weight_sum = sum(
        (float(group["weight"]) for group in minced_meat_batch_groups.values())
    )

    last_row = len([key for key, val in data.items() if val["used_weight"]]) + 3
    minced_row = last_row + 2
    minced_last_row = minced_row + len(minced_meat_batch_groups.keys()) + 3
    raw_meat_batches = RawMeatBatch.objects.filter(
        pk__in=[
            int(key.split("---")[2]) for key, val in data.items() if val["used_weight"]
        ]
    ).all()

    second_minced_meats = SecondMincedMeat.objects.filter(
        created_at__gt=start, created_at__lt=end
    )
    weight_all_second_minced_meat = sum([i.weight for i in second_minced_meats])

    for i, raw_meat_batch in enumerate(raw_meat_batches, start=1):
        meat_weight = sum(
            [
                float(val["used_weight"])
                for key, val in data.items()
                if str(raw_meat_batch.pk) in key
            ]
        )
        weight_sum += meat_weight

        meterial_params = RawMaterialParams.objects.filter(
            raw_material_id=raw_meat_batch.raw_material.pk
        ).first()
        i_tb = i + 2
        ws[f"A{i_tb}"].value = i
        ws[f"D{i_tb}"].value = raw_meat_batch.company.name
        ws[f"E{i_tb}"].value = raw_meat_batch.raw_material.name
        ws[f"F{i_tb}"].value = raw_meat_batch.raw_material.name

        separation_status = raw_meat_batch.statuses.filter(
            status__codename="laboratory_analyze"
        ).first()
        ws[f"G{i_tb}"].number_format = numbers.FORMAT_NUMBER_00
        ws[f"H{i_tb}"].number_format = numbers.FORMAT_NUMBER_00

        old = meat_weight

        ws[f"M{i_tb}"].value = (
            "" if raw_meat_batch.condition != "frozen" else meterial_params.defrost
        )
        ws[f"G{i_tb}"].value = old

        ws[f"H{i_tb}"].value = (
            old - (old * meterial_params.defrost) / 100
            if meterial_params.defrost and raw_meat_batch.condition == "frozen"
            else old
        )
        ws[f"C{i_tb}"].value = meterial_params.second_minced_meat_exit
        ws[f"G{i_tb}"].value = old
        ws[f"N{i_tb}"].value = ws[f"H{i_tb}"].value * (ws[f"C{i_tb}"].value / 100)

        if raw_meat_batch.condition == "chilled":
            ws[f"J{i_tb}"].value = "X"
        elif raw_meat_batch.condition == "frozen":
            ws[f"I{i_tb}"].value = "X"

        if separation_status:
            separator_name_mapping = {"POSS": 1, "AM2C": 2}
            ws[f"K{i_tb}"].value = separator_name_mapping[
                separation_status.additional_data["separator_name"]
            ]
            ws[f"L{i_tb}"].value = separation_status.additional_data["separator_mode"]
        i += 1
    ws.merge_cells(f"I{last_row}:M{last_row}")
    ws.merge_cells(f"A{last_row}:F{last_row}")
    ws[f"A{last_row}"].value = "Итого, кг"
    ws[f"G{last_row}"].value = weight_sum
    ws[f"G{last_row}"].number_format = numbers.FORMAT_NUMBER_00

    ws[f"I{last_row}"].value = "Итого(МКО), кг"
    ws[f"N{last_row}"].value = sum([ws[f"N{i}"].value for i in range(3, last_row)])
    ws[f"N{last_row}"].number_format = numbers.FORMAT_NUMBER_00
    if raw_meat_batches:
        ws.merge_cells(f"B3:B{len(raw_meat_batches) + 2}")
        ws["B3"].value = "прод"
    ws.merge_cells("A1:B2")
    ws["A1"].value = "№"
    ws.merge_cells("D1:F1")
    if end - start < timedelta(days=1):
        ws["D1"].value = start.strftime(date_format)
    else:
        ws["D1"].value = f"{end.strftime(date_format)}"

    ws.merge_cells("C1:C2")
    ws["C1"].value = "%Выхода"
    ws["D2"].value = "Производитель"
    ws["E2"].value = "Тип"
    ws["F2"].value = "Наименование"
    ws.merge_cells("G1:H1")
    ws["G1"].value = "Масса, кг."
    ws["G2"].value = "до дефроста"
    ws["H2"].value = "после дефроста"
    ws.merge_cells("I1:I2")
    ws["I1"].value = "ЗАМ"
    ws.merge_cells("J1:J2")
    ws["J1"].value = "ОХЛ"
    ws.merge_cells("K1:K2")
    ws["K1"].value = "Сепаратор"
    ws.merge_cells("L1:L2")
    ws["L1"].value = "режим\nсепарации"
    ws.merge_cells("M1:M2")
    ws["M1"].value = "% потери"

    ws.merge_cells("N1:N2")
    ws["N1"].value = "В В С"

    ws.merge_cells(f"A{minced_row}:E{minced_row}")
    ws[f"A{minced_row}"].value = "Выход готовой продукции"
    ws.merge_cells(f"A{minced_row + 1}:B{minced_row + 2}")
    ws[f"A{minced_row + 1}"].value = "№"
    ws.merge_cells(f"C{minced_row + 1}:C{minced_row + 2}")
    ws.merge_cells(f"D{minced_row + 1}:E{minced_row + 1}")
    ws[f"D{minced_row + 1}"].value = "Масса, кг"
    ws[f"D{minced_row + 2}"].value = "ММО"
    ws[f"E{minced_row + 2}"].value = "МКО"
    for i, (key, val) in enumerate(minced_meat_batch_groups.items(), 1):
        ws[f"C{minced_row + 3 + i - 1}"].value = key
        ws[f"D{minced_row + 3 + i - 1}"].value = val["weight"]
        ws[f"A{minced_row + 3 + i - 1}"].value = i

    ws[f"E{minced_row + 3}"].value = weight_all_second_minced_meat
    ws[f"E{minced_last_row}"].value = weight_all_second_minced_meat
    ws.merge_cells(f"E{minced_row + 1}:E{minced_row + 1}")
    if minced_meat_batch_groups:
        ws.merge_cells(
            f"B{minced_row + 3}:B{minced_row + len(minced_meat_batch_groups.keys()) + 2}"
        )
        ws.merge_cells(
            f"E{minced_row + 3}:E{minced_row + len(minced_meat_batch_groups.keys()) + 2}"
        )
        ws[f"B{minced_row + 3}"].value = "прод"

    ws.merge_cells(f"A{minced_last_row}:C{minced_last_row + 2}")
    ws[f"A{minced_last_row}"] = "ИТОГО, КГ"
    ws[f"D{minced_last_row}"].value = minced_weight_sum
    ws[f"E{minced_last_row}"].value = weight_all_second_minced_meat
    ws[f"D{minced_last_row + 1}"].value = (
        (minced_weight_sum / weight_sum) * 100 if weight_sum else ""
    )
    ws[f"E{minced_last_row + 1}"].value = (
        (weight_all_second_minced_meat / weight_sum) * 100 if weight_sum else ""
    )
    ws[f"D{minced_last_row}"].number_format = numbers.FORMAT_NUMBER_00
    ws.merge_cells(f"D{minced_last_row + 2}:E{minced_last_row + 2}")
    ws[f"D{minced_last_row + 2}"].value = minced_weight_sum
    ws[f"D{minced_last_row + 2}"].number_format = numbers.FORMAT_NUMBER_00

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    ws.column_dimensions["A"].width = 5

    if raw_meat_batches:
        ws["B3"].alignment = Alignment(
            textRotation=90, horizontal="center", vertical="center"
        )
        ws["B3"].font = Font(bold=True, size=14)
        ws.column_dimensions["B"].width = 5

    if minced_meat_batch_groups:
        ws[f"B{minced_row + 3}"].alignment = Alignment(
            textRotation=90, horizontal="center", vertical="center"
        )
        ws[f"B{minced_row + 3}"].font = Font(bold=True, size=14)
        ws.column_dimensions["B"].width = 5

    ws["D1"].font = Font(bold=True, size=16)

    ws["I1"].alignment = Alignment(
        textRotation=90, horizontal="center", vertical="center"
    )
    ws["I1"].font = Font(bold=True, size=14)
    ws.column_dimensions["L"].width = 8

    ws["J1"].alignment = Alignment(
        textRotation=90, horizontal="center", vertical="center"
    )
    ws["J1"].font = Font(bold=True, size=14)
    ws.column_dimensions["I"].width = 7

    ws["K1"].alignment = Alignment(
        textRotation=90, horizontal="center", vertical="center"
    )
    ws["K1"].font = Font(bold=True, size=10)
    ws.column_dimensions["J"].width = 7

    ws["L1"].alignment = Alignment(
        textRotation=90, horizontal="center", vertical="center", wrap_text=True
    )
    ws["L1"].font = Font(bold=True, size=10)
    ws.column_dimensions["K"].width = 7

    ws["M1"].alignment = Alignment(
        textRotation=90, horizontal="center", vertical="center", wrap_text=True
    )
    ws["M1"].font = Font(bold=True, size=10)
    ws.column_dimensions["K"].width = 7

    ws[f"A{last_row}"].font = Font(bold=True, size=16)

    ws[f"A{minced_last_row}"].font = Font(bold=True, size=16)
    ws[f"A{minced_last_row}"].alignment = Alignment(
        horizontal="center", vertical="center"
    )

    set_global_border(ws, THIN_BORDER)

    set_border_by_range(ws, f"A{last_row + 1}:M{last_row + 1}", None)
    set_border_by_range(
        ws,
        f"F{minced_row}:M{minced_row + len(minced_meat_batch_groups.keys()) + 6}",
        None,
    )
    ws[f"H{last_row}"].border = None
    ws[f"I{last_row}"].border = None
    ws[f"J{last_row}"].border = None
    ws[f"K{last_row}"].border = None
    ws[f"L{last_row}"].border = None
    beautify_columns(ws, font_size=12)
    return BytesIO(save_virtual_workbook(wb))


def get_recipes(shipments):
    recipes = list()
    for shipment in shipments:
        shipment_pallets = shipment.pallets.all()
        if shipment_pallets:
            shipment_pallet = shipment_pallets[0]
            recipes.append(
                shipment_pallet.minced_meat_batch_mix.minced_meat_batch.recipe
            )
    return recipes


def get_minced_meat_batches(shipments):
    minced_meat_batches = list()
    for shipment in shipments:
        shipment_pallets = shipment.pallets.all()
        if shipment_pallets:
            for pallet in shipment_pallets:
                minced_meat_batches.append(
                    pallet.minced_meat_batch_mix.minced_meat_batch
                )
    return list(set(minced_meat_batches))


def generate_traceability_datatable(
    start: Optional[datetime] = None, end: Optional[datetime] = None
) -> BytesIO:
    conditions = {"chilled": "Охлажденное", "frozen": "Замороженое"}

    now = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    if not start:
        start = now - timedelta(days=365)

    if not end:
        end = now

    shipments = Shipment.get_by_date_range(start, end)
    minced_meat_batches = get_minced_meat_batches(shipments)

    wb = Workbook()
    ws = wb.active
    index = 3
    ws.title = "Прослеживаемости"
    for i, minced_meat_batch in enumerate(minced_meat_batches, start=1):
        ws[f"A{index}"].value = "ID сырья"
        ws[f"B{index}"].value = "Производитель"
        ws[f"C{index}"].value = "Тип сырья"
        ws[f"D{index}"].value = "Общая масса"
        ws[f"E{index}"].value = "Кондиция"
        ws[f"F{index}"].value = "Сепаратор"
        ws[f"G{index}"].value = "Режим"

        ws[f"A{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{index}"].alignment = Alignment(
            textRotation=45, horizontal="left", vertical="center"
        )
        ws[f"C{index}"].alignment = Alignment(
            textRotation=45, horizontal="left", vertical="center"
        )
        ws[f"D{index}"].alignment = Alignment(
            textRotation=45, horizontal="left", vertical="center"
        )
        ws[f"E{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"G{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"H{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"H{index + 2}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"H{index + 3}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        ws[f"A{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"B{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"C{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"D{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"E{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"F{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"G{index}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")

        number_mix = MincedMeatBatchMix.objects.filter(
            minced_meat_batch=minced_meat_batch
        )
        if number_mix:
            for mix_index in range(len(number_mix)):
                ws.cell(row=index + 4, column=mix_index + 8).value = mix_index + 1
                ws.cell(row=index + 4, column=mix_index + 8).fill = PatternFill(
                    start_color="AEAFF2", fill_type="solid"
                )

        minced_meat_batches_raw_meat_batches = (
            MincedMeatBatchRawMeatBatch.objects.filter(
                minced_meat_batch=minced_meat_batch
            ).all()
        )
        if minced_meat_batches_raw_meat_batches:
            raw_meat_batch_index = index + 5
            for u, minced_meat_batches_raw_meat_batch in enumerate(
                minced_meat_batches_raw_meat_batches, start=1
            ):
                ws[
                    f"A{raw_meat_batch_index}"
                ].value = (
                    minced_meat_batches_raw_meat_batch.raw_meat_batch.production_id
                )
                ws[
                    f"B{raw_meat_batch_index}"
                ].value = minced_meat_batches_raw_meat_batch.raw_meat_batch.organization
                ws[
                    f"C{raw_meat_batch_index}"
                ].value = (
                    minced_meat_batches_raw_meat_batch.raw_meat_batch.raw_material.name
                )
                ws[
                    f"D{raw_meat_batch_index}"
                ].value = minced_meat_batches_raw_meat_batch.weight
                ws[f"E{raw_meat_batch_index}"].value = conditions[
                    minced_meat_batches_raw_meat_batch.raw_meat_batch.condition
                ]
                if u % 2 == 0:
                    ws[f"A{raw_meat_batch_index}"].fill = PatternFill(
                        start_color="C5D0E6", fill_type="solid"
                    )
                    ws[f"B{raw_meat_batch_index}"].fill = PatternFill(
                        start_color="C5D0E6", fill_type="solid"
                    )
                    ws[f"C{raw_meat_batch_index}"].fill = PatternFill(
                        start_color="C5D0E6", fill_type="solid"
                    )
                    ws[f"D{raw_meat_batch_index}"].fill = PatternFill(
                        start_color="C5D0E6", fill_type="solid"
                    )
                    ws[f"E{raw_meat_batch_index}"].fill = PatternFill(
                        start_color="C5D0E6", fill_type="solid"
                    )

                laboratory_status = RawMeatBatchStatus.objects.filter(
                    raw_meat_batch=minced_meat_batches_raw_meat_batch.raw_meat_batch,
                    status__codename="laboratory_analyze",
                ).first()
                if laboratory_status:
                    ws[
                        f"F{raw_meat_batch_index}"
                    ].value = laboratory_status.additional_data["separator_name"]
                    ws[
                        f"G{raw_meat_batch_index}"
                    ].value = laboratory_status.additional_data["separator_mode"]

                minced_meat_batch_mixes = MincedMeatBatchMix.objects.filter(
                    minced_meat_batch=minced_meat_batch
                ).all()
                for j, mmbm in enumerate(minced_meat_batch_mixes, start=8):
                    ws.cell(row=raw_meat_batch_index, column=j).value = sum(
                        (
                            minced_meat_batch_raw_meat_batch.weight
                            for minced_meat_batch_raw_meat_batch in minced_meat_batches_raw_meat_batches
                        )
                    )

                raw_meat_batch_index += 1

            ws.merge_cells(f"A{index}:A{index + 4}")
            ws.merge_cells(f"B{index}:B{index + 4}")
            ws.merge_cells(f"C{index}:C{index + 4}")
            ws.merge_cells(f"D{index}:D{index + 4}")
            ws.merge_cells(f"E{index}:E{index + 4}")
            ws.merge_cells(f"F{index}:F{index + 4}")
            ws.merge_cells(f"G{index}:G{index + 4}")

            p_index = index
            index = raw_meat_batch_index + 2
            if number_mix:
                ws[f"H{p_index}"].value = f"Рецепт № {minced_meat_batch.recipe.id}"
                ws[f"H{p_index + 2}"].value = f"{minced_meat_batch.recipe.name}"
                ws[f"H{p_index + 3}"].value = "№ Замеса"
                ws[f"H{p_index}"].fill = PatternFill(
                    start_color="C5D0E6", fill_type="solid"
                )
                ws[f"H{p_index + 2}"].fill = PatternFill(
                    start_color="AEAFF2", fill_type="solid"
                )
                ws[f"H{p_index + 3}"].fill = PatternFill(
                    start_color="C5D0E6", fill_type="solid"
                )
                ws.merge_cells(
                    start_row=p_index,
                    start_column=8,
                    end_row=p_index + 1,
                    end_column=len(number_mix) + 7,
                )
                ws.merge_cells(
                    start_row=p_index + 2,
                    start_column=8,
                    end_row=p_index + 2,
                    end_column=len(number_mix) + 7,
                )
                ws.merge_cells(
                    start_row=p_index + 3,
                    start_column=8,
                    end_row=p_index + 3,
                    end_column=len(number_mix) + 7,
                )
                ws.cell(
                    row=raw_meat_batch_index, column=7
                ).value = "Номер\n паллеты ГП\n после ШЗ"
                ws[f"G{raw_meat_batch_index}"].alignment = Alignment(
                    textRotation=90, horizontal="center", vertical="center"
                )
                max_mix_pallet = 0
                for j, mix in enumerate(number_mix, start=1):
                    ws.cell(row=raw_meat_batch_index, column=j + 7).value = j
                    ws.cell(row=raw_meat_batch_index, column=j + 7).fill = PatternFill(
                        start_color="AEAFF2", fill_type="solid"
                    )

                    status = mix.statuses.filter(status__codename="palletizing").first()
                    if status and status.additional_data.get("pallets"):
                        pallets = status.additional_data["pallets"]
                        if len(pallets) > max_mix_pallet:
                            max_mix_pallet = len(pallets)

                        for z, pallet in enumerate(
                            pallets, start=raw_meat_batch_index + 1
                        ):
                            ws.cell(row=z, column=j + 7).value = pallet["number"]
                index = raw_meat_batch_index + len(number_mix) + 3

                for j, mix in enumerate(number_mix):
                    for z in range(max_mix_pallet):
                        ws.cell(
                            row=z + raw_meat_batch_index + 1, column=j + 8
                        ).fill = PatternFill(start_color="5BC57B", fill_type="solid")

                ws.merge_cells(
                    f"G{raw_meat_batch_index}:G{raw_meat_batch_index + max_mix_pallet}"
                )

        ws.cell(
            row=index, column=1
        ).value = f"Партия ММО № {minced_meat_batch.recipe.id}"
        ws.cell(row=index + 1, column=1).value = "ID паллеты готовой продукции"
        ws.cell(row=index + 1, column=2).value = "Заказчик"
        ws.cell(row=index + 1, column=3).value = "Масса нетто"
        ws.cell(row=index + 1, column=4).value = "Масса брутто"
        ws.cell(row=index + 1, column=5).value = "Масса паллета"
        ws.cell(row=index + 1, column=6).value = "Масса упаковки"
        ws.cell(row=index + 1, column=7).value = "Кам. \nШоковой \nЗаморозки."
        ws.cell(
            row=index + 1, column=8
        ).value = "Температура \nпри поступлении \nна склад"
        ws.merge_cells(f"A{index}:H{index}")

        ws[f"A{index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"B{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"C{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"D{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"E{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"F{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"G{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws[f"H{index + 1}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )

        ws[f"A{index}"].fill = PatternFill(start_color="C5D0E6", fill_type="solid")
        ws[f"A{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"B{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"C{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"D{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"E{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"F{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"G{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")
        ws[f"H{index + 1}"].fill = PatternFill(start_color="AEAFF2", fill_type="solid")

        index += 2
        minced_meat_batch_mixes = MincedMeatBatchMix.objects.filter(
            minced_meat_batch=minced_meat_batch
        )
        fill_row = False
        for mix in minced_meat_batch_mixes:
            status = mix.statuses.filter(status__codename="palletizing").first()
            if status and status.additional_data.get("pallets"):
                for u, pallet in enumerate(status.additional_data["pallets"], start=1):
                    s_pallet = ShipmentPallet.objects.filter(
                        minced_meat_batch_mix=mix,
                        number=pallet["number"],
                        weight=pallet["netto"],
                    ).first()
                    customer = s_pallet.shipment.customer if s_pallet else ""
                    ws.cell(
                        row=index, column=1
                    ).value = f"{str(minced_meat_batch.recipe.id).split('/')[-1]}/{mix.id}/{pallet['number']}"
                    ws.cell(row=index, column=1).alignment = Alignment(
                        horizontal="left"
                    )
                    ws.cell(row=index, column=1).fill = PatternFill(
                        start_color="5BC57B", fill_type="solid"
                    )
                    ws.cell(row=index, column=2).value = customer
                    ws.cell(row=index, column=3).value = pallet["netto"]
                    ws.cell(row=index, column=4).value = pallet["brutto"]
                    ws.cell(row=index, column=5).value = pallet["pallet_pallet_weight"]
                    ws.cell(row=index, column=6).value = pallet["pallet_package_weight"]
                    ws.cell(row=index, column=7).value = pallet["frost_camera"]
                    ws.cell(row=index, column=8).value = pallet["temperature"]
                    if fill_row:
                        ws[f"B{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"C{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"D{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"E{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"F{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"G{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        ws[f"H{index}"].fill = PatternFill(
                            start_color="C5D0E6", fill_type="solid"
                        )
                        fill_row = False
                    else:
                        fill_row = True
                    index += 1
        index += 4
    beautify_columns(ws, font_size=14, center=False)
    set_global_border(ws, THIN_BORDER)

    return BytesIO(save_virtual_workbook(wb))


def fill_wws(table) -> BytesIO:
    wb = openpyxl.load_workbook(table)
    ws = wb.active
    rows = ws.max_row

    wss_total = 0
    last_row = 0
    for i in range(1, rows + 1):
        try:
            percent_cell = float(ws[f"C{i}"].value)
            after_defrosting_cell = float(ws[f"H{i}"].value)
            wss_value = after_defrosting_cell / 100 * percent_cell
            ws[f"M{i}"] = wss_value
            wss_total += wss_value
            last_row = i
        except Exception:
            pass
    ws[f"M{last_row + 1}"] = wss_total

    return BytesIO(save_virtual_workbook(wb))
